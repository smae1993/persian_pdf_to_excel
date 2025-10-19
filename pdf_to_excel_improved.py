#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Persian PDF to Excel Converter
Extracts tables from Persian/Arabic PDF files and converts them to Excel format.

Features:
- Handles RTL (Right-to-Left) text properly
- Preserves Persian/Arabic text direction
- Keeps numbers in correct order
- Auto-detects table headers
- Creates properly formatted Excel files with RTL support

Usage:
    python pdf_to_excel_improved.py input.pdf [output.xlsx]

Requirements:
    pip install pdfplumber openpyxl
"""

import sys
import os
import re
import argparse
try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Required packages not installed: {e}")
    print("Please install required packages:")
    print("pip install pdfplumber openpyxl")
    sys.exit(1)

def fix_persian_text(text):
    """Fix Persian text by reversing it, but keep numbers in correct order"""
    if not text or not isinstance(text, str):
        return text
    
    # Check if text is purely numeric (including Persian digits, commas, slashes, colons, dashes)
    if re.match(r'^[\d\u06F0-\u06F9,:/\-\.\s]+$', text):
        # It's just numbers/dates - don't reverse
        return text
    
    # Check if text contains Persian/Arabic characters
    persian_chars = sum(1 for c in text if '\u0600' <= c <= '\u06FF' or '\uFB50' <= c <= '\uFDFF' or '\uFE70' <= c <= '\uFEFF')
    
    if persian_chars > 0:
        # Split by spaces to handle mixed content
        parts = text.split()
        result_parts = []
        
        for part in parts:
            # Check if this part is numeric
            if re.match(r'^[\d\u06F0-\u06F9,:/\-\.]+$', part):
                # It's a number or date - don't reverse it
                result_parts.append(part)
            else:
                # It's text - reverse it
                result_parts.append(part[::-1])
        
        # Reverse the order of parts for RTL, but each number stays correct
        return ' '.join(reversed(result_parts))
    
    return text

def extract_tables_from_pdf(pdf_path, min_table_rows=3, min_columns=5):
    """
    Extract all tables from PDF
    
    Args:
        pdf_path: Path to PDF file
        min_table_rows: Minimum number of rows to consider a table valid
        min_columns: Minimum number of columns for a valid data row
        
    Returns:
        List of rows extracted from tables
    """
    all_data = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"PDF has {len(pdf.pages)} pages")
            
            for page_num, page in enumerate(pdf.pages, 1):
                print(f"Processing page {page_num}...")
                
                # Extract tables from the page
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        # Skip very small tables (header info)
                        if len(table) < min_table_rows:
                            continue
                        
                        # Process each row
                        for row_idx, row in enumerate(table):
                            if row and len(row) >= min_columns:  # Valid data row
                                # Clean each cell
                                clean_row = []
                                for cell in row:
                                    if cell:
                                        # Clean whitespace and normalize
                                        cell_text = str(cell).strip()
                                        cell_text = ' '.join(cell_text.split())
                                        # Fix Persian text direction
                                        cell_text = fix_persian_text(cell_text)
                                        clean_row.append(cell_text)
                                    else:
                                        clean_row.append("")
                                
                                # Skip empty rows
                                if any(clean_row):
                                    all_data.append(clean_row)
    
    except Exception as e:
        print(f"Error processing PDF: {e}")
        import traceback
        traceback.print_exc()
        return []
    
    return all_data

def create_excel_with_persian_headers(data_rows, output_path, sheet_name="Data", font_name='Arial'):
    """
    Create Excel file with proper Persian/Arabic text support
    
    Args:
        data_rows: List of rows to write to Excel
        output_path: Output Excel file path
        sheet_name: Name for the worksheet
        font_name: Font to use (default: Arial, can use 'B Nazanin' if available)
        
    Returns:
        True if successful, False otherwise
    """
    if not data_rows:
        print("No data to export")
        return False
    
    try:
        # Find max columns
        max_cols = max(len(row) for row in data_rows)
        
        # Pad rows to have same number of columns
        padded_data = []
        for row in data_rows:
            padded_row = row + [''] * (max_cols - len(row))
            padded_data.append(padded_row[:max_cols])
        
        # Reverse columns for RTL layout (so first column appears on the right)
        reversed_data = []
        for row in padded_data:
            reversed_data.append(list(reversed(row)))
        padded_data = reversed_data
        
        # Don't skip any rows - use the first row from PDF as header
        start_idx = 0
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit
        
        # Set up styles
        header_font = Font(name=font_name, size=12, bold=True)
        data_font = Font(name=font_name, size=11)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Use first row from data as headers
        if padded_data:
            headers_from_pdf = padded_data[0]
            start_idx = 1  # Start data from second row
            
            # Add headers from PDF
            for col_idx, header in enumerate(headers_from_pdf, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
        
        # Add data rows
        for row_idx, row_data in enumerate(padded_data[start_idx:], 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = data_font
                # Numbers should be center-aligned, text should be right-aligned
                if cell_value and re.match(r'^[\d\u06F0-\u06F9,:/\-\.]+$', str(cell_value).strip()):
                    cell.alignment = center_alignment
                else:
                    cell.alignment = right_alignment
                cell.border = border
        
        # Auto-adjust column widths
        for col_idx in range(1, max_cols + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if row.value:
                        cell_length = len(str(row.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with min and max limits
            adjusted_width = min(max(max_length * 1.3, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights
        ws.row_dimensions[1].height = 25  # Header row
        for row_idx in range(2, len(padded_data) + 2):
            ws.row_dimensions[row_idx].height = 20
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Enable RTL mode for the sheet
        ws.sheet_view.rightToLeft = True
        
        # Save workbook
        wb.save(output_path)
        print(f"✓ Excel file created successfully: {output_path}")
        print(f"  Total rows: {len(padded_data)}, Total columns: {max_cols}")
        return True
        
    except Exception as e:
        print(f"✗ Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function with command line argument support"""
    parser = argparse.ArgumentParser(
        description='Convert Persian/Arabic PDF tables to Excel format with proper RTL support.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python pdf_to_excel_improved.py input.pdf
  python pdf_to_excel_improved.py input.pdf output.xlsx
  python pdf_to_excel_improved.py input.pdf -o output.xlsx -s "My Data"
        """
    )
    
    parser.add_argument('input_pdf', help='Input PDF file path')
    parser.add_argument('output_excel', nargs='?', help='Output Excel file path (optional)')
    parser.add_argument('-o', '--output', dest='output_file', help='Output Excel file path')
    parser.add_argument('-s', '--sheet', default='Data', help='Sheet name (default: Data)')
    parser.add_argument('-f', '--font', default='Arial', help='Font name (default: Arial)')
    
    args = parser.parse_args()
    
    # Determine input and output paths
    pdf_path = args.input_pdf
    
    # Output path priority: -o flag > positional argument > auto-generated
    if args.output_file:
        excel_path = args.output_file
    elif args.output_excel:
        excel_path = args.output_excel
    else:
        # Auto-generate output filename
        base_name = os.path.splitext(pdf_path)[0]
        excel_path = f"{base_name}_converted.xlsx"
    
    # Validate input file
    if not os.path.exists(pdf_path):
        print(f"✗ Error: PDF file not found: {pdf_path}")
        sys.exit(1)
    
    if not pdf_path.lower().endswith('.pdf'):
        print(f"✗ Error: Input file must be a PDF file")
        sys.exit(1)
    
    print("=" * 60)
    print("Persian PDF to Excel Converter")
    print("=" * 60)
    print(f"Input PDF:     {pdf_path}")
    print(f"Output Excel:  {excel_path}")
    print(f"Sheet name:    {args.sheet}")
    print("=" * 60)
    print()
    
    # Extract tables
    print("Starting PDF processing...")
    data_rows = extract_tables_from_pdf(pdf_path)
    
    if not data_rows:
        print("✗ No data extracted from PDF")
        sys.exit(1)
    
    print(f"✓ Extracted {len(data_rows)} rows from PDF")
    print()
    
    # Create Excel file
    if create_excel_with_persian_headers(data_rows, excel_path, args.sheet, args.font):
        print()
        print("=" * 60)
        print("✓ Conversion completed successfully!")
        print("=" * 60)
    else:
        print()
        print("=" * 60)
        print("✗ Failed to create Excel file")
        print("=" * 60)
        sys.exit(1)


if __name__ == "__main__":
    # Check if running with arguments
    if len(sys.argv) > 1:
        main()
    else:
        # No arguments - show usage
        print("Persian PDF to Excel Converter")
        print()
        print("Usage:")
        print("  python pdf_to_excel_improved.py <input.pdf> [output.xlsx]")
        print()
        print("For more options, use --help")
        sys.exit(1)
